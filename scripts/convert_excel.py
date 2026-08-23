#!/usr/bin/env python3
"""
Convert the 4 UCAT Excel question banks to JSON files for use by the API routes.
Run: python3 scripts/convert_excel.py
Outputs: lib/data/vr-bank.json, lib/data/qr-bank.json, lib/data/sjt-bank.json, lib/data/dm-bank.json
"""

import json
import re
import os
import openpyxl
from pathlib import Path

DOWNLOADS = Path.home() / "Downloads"
OUT_DIR = Path(__file__).parent.parent / "lib" / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

def cell(row, idx):
    """Get cell value, returning '' for None."""
    if idx >= len(row):
        return ""
    v = row[idx].value
    return v if v is not None else ""

def str_cell(row, idx):
    return str(cell(row, idx)).strip()

# ─────────────────────────────────────────────────────────────────────────────
# VR
# ─────────────────────────────────────────────────────────────────────────────
def convert_vr():
    print("Converting VR...")
    wb = openpyxl.load_workbook(DOWNLOADS / "UCAT VR — 35 Passages + 1120 Question Bank.xlsx",
                                read_only=True, data_only=True)

    # Passages
    passages = []
    pmap = {}
    ws = wb["Passages"]
    rows = list(ws.rows)
    for row in rows[1:]:  # skip header
        pid  = str_cell(row, 0)
        if not pid:
            continue
        p = {
            "id":      pid,
            "title":   str_cell(row, 1),
            "topic":   str_cell(row, 2),
            "passage": str_cell(row, 3),
        }
        passages.append(p)
        pmap[pid] = p

    # Questions
    questions = []
    ws = wb["Questions"]
    rows = list(ws.rows)
    # Header: Question ID, Passage ID, Passage Title, Format, Set Difficulty, Set ID, Set Name,
    #         Question Within Set, Subtype, Question / Statement, Option A-D, Correct Answer,
    #         Highlight 1-3, Full Walkthrough, Why A-D, Why True/False/Can't Tell
    for row in rows[1:]:
        qid = str_cell(row, 0)
        if not qid:
            continue
        pid_raw = cell(row, 1)
        pid = f"P{int(float(pid_raw)):02d}" if pid_raw != "" else ""
        fmt  = str_cell(row, 3)   # MCQ or TFCT
        diff = str_cell(row, 4)   # Bronze/Silver/Gold/Diamond
        set_id = str_cell(row, 5)
        within_set = str_cell(row, 7)
        subtype = str_cell(row, 8)
        question = str_cell(row, 9)
        opt_a = str_cell(row, 10)
        opt_b = str_cell(row, 11)
        opt_c = str_cell(row, 12)
        opt_d = str_cell(row, 13)
        correct = str_cell(row, 14)  # A/B/C/D or True/False/Can't Tell
        h1 = str_cell(row, 15)
        h2 = str_cell(row, 16)
        h3 = str_cell(row, 17)
        walkthrough = str_cell(row, 18)
        why_a = str_cell(row, 19)
        why_b = str_cell(row, 20)
        why_c = str_cell(row, 21)
        why_d = str_cell(row, 22)
        why_t = str_cell(row, 23)
        why_f = str_cell(row, 24)
        why_ct = str_cell(row, 25)

        # Build supporting highlights
        highlights = [h for h in [h1, h2, h3] if h]
        # Build per-option explanations
        if fmt == "TFCT":
            option_explanations = {"True": why_t, "False": why_f, "Can't Tell": why_ct}
        else:
            option_explanations = {"A": why_a, "B": why_b, "C": why_c, "D": why_d}

        questions.append({
            "id": qid,
            "passage_id": pid,
            "format": fmt,
            "difficulty": diff,
            "set_id": set_id,
            "set_name": str_cell(row, 6),
            "question_within_set": within_set,
            "subtype": subtype,
            "question": question,
            "option_a": opt_a,
            "option_b": opt_b,
            "option_c": opt_c,
            "option_d": opt_d,
            "correct_answer": correct,
            "highlights": highlights,
            "walkthrough": walkthrough,
            "option_explanations": option_explanations,
        })

    result = {"passages": passages, "questions": questions}
    out = OUT_DIR / "vr-bank.json"
    out.write_text(json.dumps(result, indent=2, ensure_ascii=False))
    print(f"  VR: {len(passages)} passages, {len(questions)} questions → {out}")
    return result

# ─────────────────────────────────────────────────────────────────────────────
# QR
# ─────────────────────────────────────────────────────────────────────────────
def parse_qr_question(raw_q, raw_ans, raw_work, dataset_num, q_num, topic, difficulty):
    """Parse a QR question cell like 'How much does X? • A) 56 | B) 69 | C) 76 | D) 64'"""
    if not raw_q:
        return None

    # Split on • to get question text and options part
    parts = str(raw_q).split("•", 1)
    question_text = parts[0].strip()
    options = []
    if len(parts) > 1:
        # Options like "A) 56 | B) 69 | C) 76 | D) 64"
        option_str = parts[1].strip()
        for opt in re.split(r"\s*\|\s*", option_str):
            m = re.match(r"^([A-E])\)\s*(.+)$", opt.strip())
            if m:
                options.append(m.group(2).strip())

    # Parse answer like "D) 64" → correct index
    correct = 0
    if raw_ans:
        m = re.match(r"^([A-E])\)", str(raw_ans).strip())
        if m:
            correct = "ABCDE".index(m.group(1))

    return {
        "id": f"QR-{topic[:3].upper()}-D{dataset_num:03d}-Q{q_num}",
        "topic": topic,
        "difficulty": difficulty,
        "dataset_num": dataset_num,
        "question": question_text,
        "options": options,
        "correct": correct,
        "working": str(raw_work) if raw_work else "",
    }

def convert_qr():
    print("Converting QR...")
    wb = openpyxl.load_workbook(DOWNLOADS / "UCAT QR Question Bank — 72 Questions per Topic.xlsx",
                                read_only=True, data_only=True)

    datasets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.rows)
        topic = sheet_name.strip()

        for row in rows[1:]:  # skip header
            ds_num_raw = cell(row, 0)
            if ds_num_raw == "" or ds_num_raw is None:
                continue
            try:
                ds_num = int(float(ds_num_raw))
            except (ValueError, TypeError):
                continue

            diff   = str_cell(row, 1)
            title  = str_cell(row, 2)
            pres   = str_cell(row, 3)
            data   = str_cell(row, 4)

            questions = []
            for qi in range(4):
                col_q   = 5 + qi * 3
                col_ans = 6 + qi * 3
                col_wk  = 7 + qi * 3
                q = parse_qr_question(
                    cell(row, col_q), cell(row, col_ans), cell(row, col_wk),
                    ds_num, qi + 1, topic, diff
                )
                if q:
                    questions.append(q)

            if questions:
                datasets.append({
                    "id": f"QR-{topic[:3].upper()}-D{ds_num:03d}",
                    "topic": topic,
                    "difficulty": diff,
                    "title": title,
                    "presentation": pres,
                    "data": data,
                    "questions": questions,
                })

    out = OUT_DIR / "qr-bank.json"
    out.write_text(json.dumps(datasets, indent=2, ensure_ascii=False))
    print(f"  QR: {len(datasets)} datasets → {out}")
    return datasets

# ─────────────────────────────────────────────────────────────────────────────
# SJT
# ─────────────────────────────────────────────────────────────────────────────
def convert_sjt():
    print("Converting SJT...")
    wb = openpyxl.load_workbook(DOWNLOADS / "UCAT SJT Question Bank — Appropriateness, Importance & Most-Least.xlsx",
                                read_only=True, data_only=True)

    all_questions = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Category Framework":
            continue
        ws = wb[sheet_name]
        rows = list(ws.rows)
        if len(rows) < 2:
            continue

        header = [str(c.value).strip() if c.value else "" for c in rows[0]]

        # Determine format from sheet name
        name_lower = sheet_name.lower()
        if "most" in name_lower or "least" in name_lower:
            fmt = "most-least"
        elif "import" in name_lower:
            fmt = "importance"
        else:
            fmt = "appropriateness"

        # Determine difficulty from sheet name
        diff = "Bronze"
        for d in ["Bronze", "Silver", "Gold", "Diamond"]:
            if d.lower() in name_lower:
                diff = d
                break

        for row in rows[1:]:
            vals = [c.value for c in row]
            # Skip empty rows
            if not any(v for v in vals):
                continue

            scenario_id = str(vals[0]).strip() if vals[0] else ""
            if not scenario_id:
                continue

            # Difficulty from column if present
            row_diff = str(vals[1]).strip() if vals[1] else diff

            if fmt == "most-least":
                # Columns: Scenario ID, Difficulty, Scenario Title, Central Character, Scenario Text,
                # Sections Addressed, Question ID, Action A, Action B, Action C,
                # Most Appropriate, Least Appropriate, Rationale A, Rationale B, Rationale C
                q = {
                    "id": str(vals[6]).strip() if len(vals) > 6 and vals[6] else f"{scenario_id}-ML",
                    "format": "most-least",
                    "difficulty": row_diff,
                    "scenario_id": scenario_id,
                    "scenario_title": str(vals[2]).strip() if len(vals) > 2 and vals[2] else "",
                    "character": str(vals[3]).strip() if len(vals) > 3 and vals[3] else "",
                    "scenario": str(vals[4]).strip() if len(vals) > 4 and vals[4] else "",
                    "sections": str(vals[5]).strip() if len(vals) > 5 and vals[5] else "",
                    "action_a": str(vals[7]).strip() if len(vals) > 7 and vals[7] else "",
                    "action_b": str(vals[8]).strip() if len(vals) > 8 and vals[8] else "",
                    "action_c": str(vals[9]).strip() if len(vals) > 9 and vals[9] else "",
                    "most_correct": str(vals[10]).strip() if len(vals) > 10 and vals[10] else "A",
                    "least_correct": str(vals[11]).strip() if len(vals) > 11 and vals[11] else "C",
                    "rationale_a": str(vals[12]).strip() if len(vals) > 12 and vals[12] else "",
                    "rationale_b": str(vals[13]).strip() if len(vals) > 13 and vals[13] else "",
                    "rationale_c": str(vals[14]).strip() if len(vals) > 14 and vals[14] else "",
                }
            else:
                # Columns: Scenario ID, Difficulty, Scenario Title, Central Character, Scenario Text,
                # Sections Addressed, Question ID, Action/Consideration, Correct Rating,
                # Core Explanation, Why A, Why B, Why C, Why D
                q = {
                    "id": str(vals[6]).strip() if len(vals) > 6 and vals[6] else f"{scenario_id}-Q",
                    "format": fmt,
                    "difficulty": row_diff,
                    "scenario_id": scenario_id,
                    "scenario_title": str(vals[2]).strip() if len(vals) > 2 and vals[2] else "",
                    "character": str(vals[3]).strip() if len(vals) > 3 and vals[3] else "",
                    "scenario": str(vals[4]).strip() if len(vals) > 4 and vals[4] else "",
                    "sections": str(vals[5]).strip() if len(vals) > 5 and vals[5] else "",
                    "action": str(vals[7]).strip() if len(vals) > 7 and vals[7] else "",
                    "correct_rating": str(vals[8]).strip() if len(vals) > 8 and vals[8] else "",
                    "explanation": str(vals[9]).strip() if len(vals) > 9 and vals[9] else "",
                    "why_a": str(vals[10]).strip() if len(vals) > 10 and vals[10] else "",
                    "why_b": str(vals[11]).strip() if len(vals) > 11 and vals[11] else "",
                    "why_c": str(vals[12]).strip() if len(vals) > 12 and vals[12] else "",
                    "why_d": str(vals[13]).strip() if len(vals) > 13 and vals[13] else "",
                }

            all_questions.append(q)

    out = OUT_DIR / "sjt-bank.json"
    out.write_text(json.dumps(all_questions, indent=2, ensure_ascii=False))
    print(f"  SJT: {len(all_questions)} questions → {out}")
    return all_questions

# ─────────────────────────────────────────────────────────────────────────────
# DM — complex multi-format block parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_arguments_sheet(ws, difficulty):
    """Parse Arguments/Assumptions sheets (block format)."""
    rows = list(ws.rows)
    questions = []
    current_q = None
    in_options = False
    q_num = 0

    for row in rows:
        vals = [c.value for c in row]
        if not vals:
            continue
        r0 = str(vals[0]).strip() if vals[0] else ""
        r1 = str(vals[1]).strip() if len(vals) > 1 and vals[1] else ""

        # Detect question header: "QUESTION 01 — Title | BRONZE | TYPE"
        if re.match(r"^QUESTION\s+\d+", r0):
            if current_q:
                questions.append(current_q)
            q_num += 1
            m = re.match(r"^QUESTION\s+\d+\s+—\s+(.+?)\s*\|\s*\w+\s*\|\s*(.+?)$", r0)
            qtype = m.group(2).strip() if m else "Arguments"
            qtitle = m.group(1).strip() if m else r0
            current_q = {
                "id": f"DM-ARG-{difficulty[0]}-{q_num:03d}",
                "type": "Arguments",
                "question_type": qtype,
                "difficulty": difficulty,
                "title": qtitle,
                "task": "",
                "options": [],
                "answer": "",
                "walkthrough": "",
                "key_rule": "",
            }
            in_options = False

        elif current_q and r0 == "TASK":
            current_q["task"] = r1

        elif current_q and r0 == "Option":
            in_options = True

        elif current_q and in_options and r0 in ("A", "B", "C", "D"):
            is_correct = str(vals[2]).strip().upper() in ("YES", "✓") if vals[2] else False
            current_q["options"].append({
                "letter": r0,
                "text": r1,
                "is_correct": is_correct,
                "explanation": str(vals[3]).strip() if vals[3] else "",
            })

        elif current_q and r0 == "ANSWER":
            current_q["answer"] = r1
            in_options = False

        elif current_q and r0 == "WALKTHROUGH":
            current_q["walkthrough"] = r1

        elif current_q and r0 == "KEY RULE":
            current_q["key_rule"] = r1

    if current_q:
        questions.append(current_q)

    return questions


def parse_logical_sheet(ws, difficulty):
    """Parse Logical Puzzles sheets."""
    rows = list(ws.rows)
    questions = []
    current_q = None
    q_num = 0
    in_clues = False

    for row in rows:
        vals = [c.value for c in row]
        r0 = str(vals[0]).strip() if vals[0] else ""
        r1 = str(vals[1]).strip() if vals[1] else ""

        if re.match(r"^QUESTION\s+\d+", r0) and "BRONZE" in r0.upper() or re.match(r"^QUESTION\s+\d+\s+—", r0):
            if current_q:
                questions.append(current_q)
            q_num += 1
            m = re.match(r"^QUESTION\s+\d+\s+—\s+(.+?)(?:\s*\|.*)?$", r0)
            qtitle = m.group(1).strip() if m else r0
            current_q = {
                "id": f"DM-LOG-{difficulty[0]}-{q_num:03d}",
                "type": "Logical",
                "difficulty": difficulty,
                "title": qtitle,
                "scenario": "",
                "question": "",
                "options": [],
                "clues": [],
                "answer": "",
                "walkthrough": "",
                "blank_grid": "",
                "completed_grid": "",
            }
            in_clues = False

        elif current_q and r0 == "Scenario":
            current_q["scenario"] = r1

        elif current_q and r0 == "Question":
            current_q["question"] = r1

        elif current_q and re.match(r"^Option [A-D]$", r0):
            letter = r0[-1]
            is_correct = str(vals[2]).strip().upper() == "CORRECT" if vals[2] else False
            current_q["options"].append({
                "letter": letter,
                "text": r1,
                "is_correct": is_correct,
            })

        elif current_q and r0 == "CLUES":
            in_clues = True

        elif current_q and in_clues and re.match(r"^Clue\s+\d+$", r0):
            current_q["clues"].append({"clue": r1, "hint": str(vals[3]).strip() if vals[3] else ""})

        elif current_q and r0 == "BLANK GRID":
            current_q["blank_grid"] = r1
            in_clues = False

        elif current_q and r0 == "COMPLETED GRID":
            current_q["completed_grid"] = r1

        elif current_q and r0 == "ANSWER":
            current_q["answer"] = r1

        elif current_q and r0 == "KEY METHOD":
            current_q["walkthrough"] = r1

    if current_q:
        questions.append(current_q)

    return questions


def parse_probability_sheet(ws, difficulty):
    """Parse Probability sheets."""
    rows = list(ws.rows)
    questions = []
    current_q = None
    q_num = 0
    in_options = False

    for row in rows:
        vals = [c.value for c in row]
        r0 = str(vals[0]).strip() if vals[0] else ""
        r1 = str(vals[1]).strip() if vals[1] else ""

        if re.match(r"^QUESTION\s+\d+", r0):
            if current_q:
                questions.append(current_q)
            q_num += 1
            m = re.match(r"^QUESTION\s+\d+\s+—\s+(.+?)$", r0)
            qtitle = m.group(1).strip() if m else r0
            current_q = {
                "id": f"DM-PRB-{difficulty[0]}-{q_num:03d}",
                "type": "Probability",
                "difficulty": difficulty,
                "title": qtitle,
                "question": "",
                "options": [],
                "worked_solution": "",
                "answer": "",
                "key_method": "",
            }
            in_options = False

        elif current_q and r0 == "QUESTION":
            current_q["question"] = r1

        elif current_q and r0 == "Option":
            in_options = True

        elif current_q and in_options and r0 in ("A", "B", "C", "D"):
            is_correct = str(vals[2]).strip().upper() in ("YES", "✓") if vals[2] else False
            exact_val = vals[5] if len(vals) > 5 else None
            current_q["options"].append({
                "letter": r0,
                "text": r1,
                "is_correct": is_correct,
                "explanation": str(vals[3]).strip() if vals[3] else "",
                "exact_value": float(exact_val) if exact_val is not None and exact_val != "" else None,
            })

        elif current_q and r0 == "WORKED SOLUTION":
            current_q["worked_solution"] = r1
            in_options = False

        elif current_q and r0 == "ANSWER":
            current_q["answer"] = r1

        elif current_q and r0 == "KEY METHOD":
            current_q["key_method"] = r1

    if current_q:
        questions.append(current_q)

    return questions


def parse_interpreting_sheet(ws, difficulty):
    """Parse Interpreting Information sheets (YN-5 format)."""
    rows = list(ws.rows)
    questions = []
    current_q = None
    q_num = 0
    in_table = False
    in_rules = False
    in_conclusions = False
    table_headers = []
    table_rows = []
    rules = []

    for row in rows:
        vals = [c.value for c in row]
        r0 = str(vals[0]).strip() if vals[0] else ""
        r1 = str(vals[1]).strip() if vals[1] else ""

        if re.match(r"^QUESTION\s+\d+", r0):
            if current_q:
                questions.append(current_q)
            q_num += 1
            m = re.match(r"^QUESTION\s+\d+\s+—\s+(.+?)(?:\s*\|.*)?$", r0)
            qtitle = m.group(1).strip() if m else r0
            current_q = {
                "id": f"DM-INT-{difficulty[0]}-{q_num:03d}",
                "type": "Interpreting",
                "difficulty": difficulty,
                "title": qtitle,
                "source_info": "",
                "table_headers": [],
                "table_rows": [],
                "rules": [],
                "conclusions": [],
            }
            in_table = False
            in_rules = False
            in_conclusions = False
            table_headers = []
            table_rows = []
            rules = []

        elif current_q and r0 == "SOURCE INFORMATION":
            current_q["source_info"] = r1
            in_table = True
            in_rules = False
            in_conclusions = False

        elif current_q and in_table:
            # Table header row — first row after SOURCE INFORMATION that has text in multiple cells
            non_empty = [str(v).strip() for v in vals if v is not None and str(v).strip()]
            if non_empty and not current_q["table_headers"] and r0 not in ("RULES", "CONCLUSIONS", "SOURCE INFORMATION", "SCORING"):
                current_q["table_headers"] = non_empty

            elif non_empty and current_q["table_headers"] and r0 not in ("RULES", "CONCLUSIONS", "SCORING"):
                current_q["table_rows"].append([str(v).strip() if v is not None else "" for v in vals[:len(current_q["table_headers"])]])

            if r0 == "RULES":
                in_table = False
                in_rules = True

        elif current_q and r0 == "RULES":
            in_rules = True
            in_table = False

        elif current_q and in_rules:
            if r0 == "CONCLUSIONS":
                in_rules = False
                in_conclusions = True
            elif r0.startswith("Rule"):
                current_q["rules"].append(r1)

        elif current_q and (r0 == "CONCLUSIONS" or in_conclusions):
            in_conclusions = True
            # Conclusion rows: #, Conclusion, Answer, Evidence, Detailed walkthrough, Key trap
            if r0 == "CONCLUSIONS" or r0 == "#":
                pass  # header
            elif r0 == "SCORING":
                in_conclusions = False
            else:
                try:
                    num = int(float(r0))
                    ans = str(vals[2]).strip().upper() if vals[2] else ""
                    current_q["conclusions"].append({
                        "num": num,
                        "statement": r1,
                        "answer": ans,  # YES or NO
                        "evidence": str(vals[3]).strip() if vals[3] else "",
                        "walkthrough": str(vals[4]).strip() if vals[4] else "",
                    })
                except (ValueError, TypeError):
                    pass

    if current_q:
        questions.append(current_q)

    return questions


def parse_venn_sheet(ws, difficulty):
    """Parse Venn Diagram sheets."""
    rows = list(ws.rows)
    questions = []
    current_q = None
    q_num = 0

    for row in rows:
        vals = [c.value for c in row]
        r0 = str(vals[0]).strip() if vals[0] else ""
        r1 = str(vals[1]).strip() if vals[1] else ""

        if re.match(r"^QUESTION\s+\d+", r0):
            if current_q:
                questions.append(current_q)
            q_num += 1
            m = re.match(r"^QUESTION\s+\d+\s+—\s+(.+?)(?:\s*\|.*)?$", r0)
            qtitle = m.group(1).strip() if m else r0
            venn_type = ""
            if "DIAGRAM GIVEN" in r0.upper():
                venn_type = "diagram-given"
            elif "DESCRIPTION" in r0.upper():
                venn_type = "description-given"
            current_q = {
                "id": f"DM-VEN-{difficulty[0]}-{q_num:03d}",
                "type": "Venn",
                "venn_type": venn_type,
                "difficulty": difficulty,
                "title": qtitle,
                "setup": "",
                "venn_diagram": "",
                "question": "",
                "options": [],
                "answer": "",
                "walkthrough": [],
            }

        elif current_q and r0 == "SETUP":
            current_q["setup"] = r1

        elif current_q and r0 == "VENN DIAGRAM":
            # The venn diagram values may be in nearby cells
            # Collect all non-empty values in this and next few rows
            pass

        elif current_q and r0 == "QUESTION":
            current_q["question"] = r1

        elif current_q and re.match(r"^Option [A-D]$", r0):
            letter = r0[-1]
            text_val = r1
            # Try to convert to number if possible
            try:
                text_val = str(int(float(r1))) if r1 else r1
            except (ValueError, TypeError):
                pass
            is_correct = str(vals[2]).strip().upper() == "CORRECT" if vals[2] else False
            current_q["options"].append({
                "letter": letter,
                "text": text_val,
                "is_correct": is_correct,
            })

        elif current_q and r0 == "WALKTHROUGH":
            pass  # steps follow

        elif current_q and re.match(r"^Step\s+\d+$", r0):
            current_q["walkthrough"].append(r1)

        elif current_q and r0 == "ANSWER":
            current_q["answer"] = r1

    if current_q:
        questions.append(current_q)

    return questions


def parse_syllogism_sheet(ws, difficulty):
    """Parse Syllogism sheets (YN-5 format)."""
    rows = list(ws.rows)
    questions = []
    current_q = None
    q_num = 0
    in_premises = False
    in_statements = False
    premises = []

    for row in rows:
        vals = [c.value for c in row]
        if not vals:
            continue
        r0 = str(vals[0]).strip() if vals[0] else ""
        r1 = str(vals[1]).strip() if len(vals) > 1 and vals[1] else ""

        if re.match(r"^QUESTION\s+\d+", r0):
            if current_q:
                questions.append(current_q)
            q_num += 1
            m = re.match(r"^QUESTION\s+\d+\s+—\s+(.+?)(?:\s*\|.*)?$", r0)
            qtitle = m.group(1).strip() if m else r0
            current_q = {
                "id": f"DM-SYL-{difficulty[0]}-{q_num:03d}",
                "type": "Syllogism",
                "difficulty": difficulty,
                "title": qtitle,
                "premises": [],
                "shorthand": [],
                "logic_map": "",
                "statements": [],
            }
            in_premises = True
            in_statements = False
            premises = []

        elif current_q and r0 == "PREMISES":
            in_premises = True
            in_statements = False

        elif current_q and in_premises:
            if r0 in ("SHORTHAND / CONSTRUCTION", "SHORTHAND"):
                in_premises = False
            elif r0 and re.match(r"^\d+\.$", r0):
                current_q["premises"].append(r1)

        elif current_q and r0 in ("SHORTHAND / CONSTRUCTION", "SHORTHAND"):
            in_premises = False

        elif current_q and r0 == "LOGIC DIAGRAM / MAP":
            current_q["logic_map"] = r1

        elif current_q and r0 == "#":
            in_statements = True

        elif current_q and in_statements:
            try:
                num = int(float(r0))
                ans_raw = str(vals[2]).strip().upper() if len(vals) > 2 and vals[2] else ""
                current_q["statements"].append({
                    "num": num,
                    "statement": r1,
                    "answer": ans_raw,
                    "step": str(vals[3]).strip() if len(vals) > 3 and vals[3] else "",
                    "conclusion": str(vals[4]).strip() if len(vals) > 4 and vals[4] else "",
                    "explanation": str(vals[5]).strip() if len(vals) > 5 and vals[5] else "",
                    "key_rule": str(vals[6]).strip() if len(vals) > 6 and vals[6] else "",
                })
            except (ValueError, TypeError):
                pass

    if current_q:
        questions.append(current_q)

    return questions


def convert_dm():
    print("Converting DM...")
    wb = openpyxl.load_workbook(DOWNLOADS / "UCAT Decision Making Question Bank.xlsx",
                                read_only=True, data_only=True)

    all_questions = []

    SHEET_PARSERS = {
        "Arguments": parse_arguments_sheet,
        "Logical": parse_logical_sheet,
        "Probability": parse_probability_sheet,
        "Interpreting": parse_interpreting_sheet,
        "Venn": parse_venn_sheet,
        "Syllogism": parse_syllogism_sheet,
    }

    for sheet_name in wb.sheetnames:
        # Determine type and difficulty from sheet name
        sn = sheet_name.strip()
        qtype = None
        diff = "Bronze"

        if "Arguments" in sn or "Assumptions" in sn:
            qtype = "Arguments"
        elif "Logical" in sn:
            qtype = "Logical"
        elif "Probability" in sn:
            qtype = "Probability"
        elif "Interpreting" in sn:
            qtype = "Interpreting"
        elif "Venn" in sn:
            qtype = "Venn"
        elif "Syllogism" in sn:
            qtype = "Syllogism"

        if qtype is None:
            continue  # Skip bank/overview sheets

        for d in ["Bronze", "Silver", "Gold", "Diamond"]:
            if d in sn:
                diff = d
                break

        parser = SHEET_PARSERS[qtype]
        ws = wb[sheet_name]
        qs = parser(ws, diff)
        print(f"  {sn}: {len(qs)} questions")
        all_questions.extend(qs)

    out = OUT_DIR / "dm-bank.json"
    out.write_text(json.dumps(all_questions, indent=2, ensure_ascii=False))
    print(f"  DM total: {len(all_questions)} question blocks → {out}")
    return all_questions


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    convert_vr()
    convert_qr()
    convert_sjt()
    convert_dm()
    print("\nDone. JSON files written to lib/data/")
